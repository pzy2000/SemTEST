import os
import subprocess
import sys
import tensorflow as tf
import tf2onnx
import torch
from onnx2torch import convert
from tqdm import tqdm

sys.path.append("../")
sys.path.append("../implementations")
from implementations.scripts.prediction.custom_objects import custom_objects

envs = "CUDA_HOME=/usr/local/cuda-10 CUDA_ROOT=/usr/local/cuda-10 LD_LIBRARY_PATH=/usr/local/cuda-10/lib64:$LD_LIBRARY_PATH PATH=/usr/local/cuda-10/bin:$PATH"
import onnxruntime as ort
import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook


def _init_input(input_shape):
    input_shape = list(input_shape)
    input_shape[0] = 10
    input_shape = tuple(input_shape)
    input = np.random.rand(*input_shape)
    return input


comet_path = r'/root/LOGS/LEMON_result'
path = comet_path
models = os.listdir(comet_path)
nan_path = []
inconsistency_path = []

for model in models:
    result = pd.read_excel(comet_path + "/" + model + "/output_coverage.xlsx")
    for i in range(result.shape[0]):
        distance = result.iloc[i, :]['distance']
        path = result.iloc[i, :]['path']
        # path = path.replace("/data1/pzy/MUTANTS/LEMON", "G:/mutants/COMET-mutants")
        if np.isnan(distance):
            nan_path.append(path)
        elif distance > 8:
            inconsistency_path.append(path)


# print(nan_path)
# print(inconsistency_path)
# exit(666)

def clear_cache():
    if os.path.exists("tmp"):
        subprocess.call("rm -rf tmp", shell=True)
    if os.path.exists("onnx"):
        subprocess.call("rm -rf onnx", shell=True)
    if not os.path.exists("tmp"):
        os.makedirs("tmp")
    if not os.path.exists("onnx"):
        os.makedirs("onnx")


clear_cache()
# if os.path.exists("Bug_log.xlsx"):
#     os.remove("Bug_log.xlsx")
# print("sleep 2s")
# time.sleep(2)
# print("awake")
if not os.path.exists("tmp"):
    os.makedirs("tmp")
# if not os.path.exists("py"):
#     os.makedirs("py")
if not os.path.exists("onnx"):
    os.makedirs("onnx")
f = open("BUG_Analysis_FAILED.txt", "w")


def _init_input(input_shape):
    input_shape = list(input_shape)
    input_shape[0] = 10
    input_shape = tuple(input_shape)
    input = np.random.rand(*input_shape)
    return input


def load_and_infer_onnx_model(model_path, input_data):
    # 加载ONNX模型
    input_data = input_data.astype(np.float32)
    try:
        session = ort.InferenceSession(model_path)
    except Exception as e:
        print(e)
        return -666
    input_name = session.get_inputs()[0].name

    # 推理
    outputs = session.run(None, {input_name: input_data})

    # 检查推理结果是否含有NaN
    has_nan = np.isnan(outputs).any()

    return has_nan


banned_model_path = ['onnx/resnet50-imagenet_origin-Edge2-NLAll4-MDtype12-LMerg27-MDtype87.onnx',
                     'onnx/resnet50-imagenet_origin-Edge2-NLAll4-MDtype12-NLAll16-NLAll32.onnx']
for filenames in tqdm(nan_path):
    # print("1111")
    # for filename in filenames:
    # print("2222")
    filename = filenames
    # print("filename", filename)
    if filename.endswith(".h5"):
        # print("33333")
        model_path = filename
        # cur_path = model_to_json(model_path)
        # union_json(cur_path, os.path.join("all_layer_info.json"))
        # try:
        # print("model_path", model_path)
        model = tf.keras.models.load_model(model_path, custom_objects=custom_objects())
        # model = tf.keras.models.load_model(model_path,)
        model_name = os.path.basename(filename)[:-3]
        # print("model_name", model_name)
        # model.save(f'tmp/{model_name}', save_format='tf')
        # status = subprocess.call(
        #     f"{envs} python -u -m tf2onnx.convert --saved-model " + f'tmp/{model_name}' + " --output onnx/" + model_name + ".onnx",
        #     shell=True)
        model_proto, _ = tf2onnx.convert.from_keras(model, opset=13, output_path="onnx/" + model_name + ".onnx")
        # print("model name", model_name)
        onnx_model_path = "onnx/" + model_name + ".onnx"
        print("onnx_model_path", onnx_model_path)
        if onnx_model_path in banned_model_path:
            continue
        try:
            torch_model = convert(onnx_model_path)
            torch_input = torch.tensor(inpu, dtype=torch.float32)
            torch_output = torch_model(torch_input)
            nan_check_torch = torch.isnan(torch_output).any()
            if_torch_nan = nan_check_torch.item()
        except Exception as e:
            if_torch_nan = -666
            print(e)
        inpu = _init_input(model.input_shape)
        try:
            onnx_has_nan = load_and_infer_onnx_model(onnx_model_path, inpu)
        except Exception as e:
            onnx_has_nan = -666
            print(e)
        print(f'onnx推理结果中含有NaN: {onnx_has_nan}')

        tf_input = tf.convert_to_tensor(inpu)

        tf_output = model(tf_input)
        nan_check_tf = tf.reduce_any(tf.math.is_nan(tf_output))
        # To evaluate this condition, you need to run it within a TensorFlow session (for TensorFlow 1.x)
        # or directly if you are using TensorFlow 2.x
        if_tf_nan = nan_check_tf.numpy()
        # print("Is there any NaN in tf_output?", if_tf_nan)
        # print("Is there any NaN in torch_output?", if_torch_nan)
        # Prepare data for DataFrame
        data = {
            'path': [str(model_path)],
            'if_tf_nan': [if_tf_nan],
            'if_torch_nan': [if_torch_nan],
            'if_onnx_nan': [onnx_has_nan]
        }
        df = pd.DataFrame(data)
        # Save the DataFrame to a file
        excel_path = 'Bug_log.xlsx'  # Define your path and file name
        if os.path.exists(excel_path):
            book = load_workbook(excel_path)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                startrow = book.worksheets[0].max_row
                df.to_excel(writer, index=False, header=False, startrow=startrow)
        else:
            # Export to Excel for the first time (file does not exist)
            df.to_excel(excel_path, index=False)
        print(f"Results exported to {excel_path}")
        # except Exception as e:
        #     print(e)
        #     f.write("model path: " + model_path + " FAILED!!!!\n")
        #     f.write(str(e))
        #     f.write("\n")
        # continue
        clear_cache()
        del model
        if if_torch_nan != -666:
            del torch_model
f.close()
